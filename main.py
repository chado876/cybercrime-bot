from openpyxl import Workbook, load_workbook


def get_q_and_a():
    question = input('Enter the question: ')
    answer = input('Enter the answer: ')
    return question, answer


def sync_to_xls(questions, answers):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'FAQS'
        newRowLocation = ws.max_row + 1
        for count, question in enumerate(questions):
            print(question)
            ws.cell(column=1, row=count+1, value=question)
            ws.cell(column=2, row=count+1, value=answers[count])
        wb.save(filename='faqs.xlsx')
        wb.close()
        print("FAQs added to spreadsheet successfully.")
    except:
        print("Could not add FAQs to spreadsheet.")


def get_faqs():
    res = ''
    questions = []
    answers = []
    while res != 'n':
        question, answer = get_q_and_a()
        questions.append(question)
        answers.append(answer)
        res = input(
            "Would you like to continue adding FAQs? 'y' for yes, 'n' for no.")

    print(questions, answers)
    sync_to_xls(questions, answers)


get_faqs()
